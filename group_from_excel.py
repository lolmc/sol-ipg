'''
This script imports data from an exported IPAM excel sheet from Solarwinds and
 converts the IP subnets into the required format for importing into the NTA IP Groups module.
The format is an xml file and follows the schema below:
<AddressGroups xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns="http://tempuri.org/IPAddressGroupsSchema.xsd">
  <AddressGroup enabled="false" description="Class A Private Addresses">
    <Range from="10.0.0.0" to="10.255.255.255" />
  </AddressGroup>
  <AddressGroup enabled="false" description="Class B Private Addresses">
    <Range from="172.16.0.0" to="172.31.255.255" />
  </AddressGroup>
  <AddressGroup enabled="false" description="Class C Private Addresses">
    <Range from="192.168.0.0" to="192.168.255.255" />
  </AddressGroup>
  <AddressGroup enabled="true" description="Guest_SSID">
    <Range from="192.168.208.1" to="192.168.211.254" />
    <Range from="192.168.204.1" to="192.168.207.254" />
    <Range from="192.168.203.1" to="192.168.203.254" />
    <Range from="192.168.202.1" to="192.168.202.254" />
    <Range from="192.168.201.1" to="192.168.201.254" />
  </AddressGroup>
 </AddressGroups>
'''
'''Individual IP addresses take the format:
  <AddressGroup enabled="true" description="Main proxy Server">
    <Range from="192.168.1.1" to="192.168.1.1" />
  </AddressGroup>
The purpose of the script is to extract the IP subnet range configured and identifier e.g. Mere-Lane Data Vlan, and 
add it as an entry to the xml file outlined in the schema.'''

# Import `os`
import os
from IPy import IP

# Retrieve current working directory (`cwd`)
cwd = os.getcwd()
cwd

# Change directory
os.chdir("C:/Users/mcbridl/PycharmProjects/nta_ip_groups")

# List all files and directories in current directory
os.listdir('.')

# Assign spreadsheet filename to `file`
file = 'Report_IPAM_-_All_Subnets.xlsx'

# Create the xls file
ipgheader = '''<?xml version="1.0" encoding="utf-8"?>
<AddressGroups xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns="http://tempuri.org/IPAddressGroupsSchema.xsd">
  <AddressGroup enabled="false" description="Class A Private Addresses">
    <Range from="10.0.0.0" to="10.255.255.255" />
  </AddressGroup>
  <AddressGroup enabled="false" description="Class B Private Addresses">
    <Range from="172.16.0.0" to="172.31.255.255" />
  </AddressGroup>
  <AddressGroup enabled="false" description="Class C Private Addresses">
    <Range from="192.168.0.0" to="192.168.255.255" />
  </AddressGroup>
    '''
ipg_end = '''</AddressGroups>'''
# set up the amended.xml file with the neccesary opening header strings ready for the dynamic appends from the script
# Create the file to hold the xml data
f = open('amended.xml', 'w')
f.write(ipgheader)
with open('amended.xml') as f:
    for line in f:
        print(line, end='')
# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('Report_IPAM_-_All_Subnets.xlsx')
ws = wb.active

# Get sheet names
# print(wb.get_sheet_names())

# Look at each row in the sheet
for row in ws.values:
    # print(row)
    # Get the value for each column in the row
    # Subnet name
    desc = row[0]
    # print (desc)
    # ip should contain the IP subnet network address in dotted decimal as a string variable
    ip = row[1]
    # cidr should contain the cidr mask notation as an integer
    cidr = row[2]
    # print out both variables
    # print(ip, cidr)
    # subnet should contain the network address/cidr mask
    subnet = IP(ip + '/' + str(cidr))
    range = (subnet.strNormal(3).split('-'))
    ipg_open = '''  <AddressGroup enabled="true" description="''' + desc + '''">\n'''
    ipg_range = '''    <Range from="''' + range[0] + '''"''' + ''' to="''' + range[1] + '''" />\n'''
    ipg_close = '''  </AddressGroup>\n'''
    ipg_end = '''</AddressGroups>'''
    f = open('amended.xml', 'a')
    f.write(ipg_open)
    f.write(ipg_range)
    f.write(ipg_close)
f = open('amended.xml', 'a')
f.write(ipg_end)
